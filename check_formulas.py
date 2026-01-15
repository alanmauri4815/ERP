import openpyxl

wb = openpyxl.load_workbook('ERP Universal.xlsx', data_only=False)

sheets_to_check = ['Productos', 'Stock', 'Inventario']

formula_report = {}

for sheet_name in sheets_to_check:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_formulas = []
        # Check first 10 rows, columns with formulas
        for row in range(1, 11):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    sheet_formulas.append({
                        'cell': cell.coordinate,
                        'formula': cell.value
                    })
        formula_report[sheet_name] = sheet_formulas

print(formula_report)
